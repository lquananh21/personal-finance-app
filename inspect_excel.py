
import openpyxl
import json

file_path = r'C:\Users\lvqua\OneDrive\Desktop\Work Life Management\01 Life\01 Finance\FINANCIAL PLANNER.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    
    summary = {}
    
    # Net worth: Read all
    if "Net worth" in sheets:
        ws = wb["Net worth"]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        summary["Net worth"] = rows

    # 2024 Financials: Read 200 rows
    if "2024 Financials" in sheets:
        ws = wb["2024 Financials"]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
            rows.append(list(row))
        summary["2024 Financials"] = rows

    # 2026 Bank accounts: Read all
    if "2026 Bank accounts" in sheets:
        ws = wb["2026 Bank accounts"]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        summary["2026 Bank accounts"] = rows
        
    with open('excel_structure.json', 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)
    print("Exported to excel_structure.json")

except Exception as e:
    with open('excel_structure.json', 'w', encoding='utf-8') as f:
        json.dump({"error": str(e)}, f)
    print(str(e))
