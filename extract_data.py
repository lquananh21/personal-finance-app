
import openpyxl
import json
import re

file_path = r'C:\Users\lvqua\OneDrive\Desktop\Work Life Management\01 Life\01 Finance\FINANCIAL PLANNER.xlsx'

def get_category(name):
    name = name.lower()
    if 'salary' in name or 'wage' in name: return 'salary'
    if 'dividend' in name: return 'investments'
    if 'pension' in name: return 'pension'
    if 'rent' in name: return 'passive'
    if 'equity' in name: return 'stocks'
    return 'other'

def extract():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {"income": [], "expenses": [], "assets": [], "liabilities": [], "investments": [], "goals": [], "budget": []}

    # 1. Net Worth (Assets/Investments)
    if "Net worth" in wb.sheetnames:
        ws = wb["Net worth"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            val = row[1]
            if name and isinstance(val, (int, float)):
                name_str = str(name).strip()
                if "TOTAL" in name_str.upper(): continue
                cat = get_category(name_str)
                item = {
                    "id": abs(hash(name_str)) % 100000,
                    "name": name_str,
                    "category": cat,
                    "currentValue": float(val),
                    "date": "2024-01-01"
                }
                # Classify as Investment or Asset
                if cat in ['stocks', 'pension', 'investments']:
                    # Investment
                    data['investments'].append({
                        "id": item['id'],
                        "name": name_str,
                        "category": cat,
                        "quantity": 1,
                        "currentPrice": float(val),
                        "costBasis": float(val) * 0.8, # Est
                        "currentValue": float(val),
                        "date": "2024-01-01"
                    })
                else:
                    # Asset
                    data['assets'].append({
                        "id": item['id'],
                        "name": name_str,
                        "type": 'Physical' if 'physical' in name_str.lower() else 'Cash',
                        "value": float(val),
                        "liquidity": 'Medium',
                        "date": "2024-01-01"
                    })

    # 2. 2024 Financials (Income/Expenses)
    if "2024 Financials" in wb.sheetnames:
        ws = wb["2024 Financials"]
        rows = list(ws.iter_rows(values_only=True))
        
        # Find Sections
        inc_start = -1
        exp_start = -1
        
        for i, row in enumerate(rows):
            if not row[1]: continue
            txt = str(row[1]).upper()
            if "INCOME" in txt and "TOTAL" not in txt: inc_start = i
            if "EXPENSES" in txt and "TOTAL" not in txt: exp_start = i
            
        # Extract Income
        if inc_start != -1:
            for i in range(inc_start + 1, len(rows)):
                row = rows[i]
                if not row[1]: continue
                desc = str(row[1])
                if "TOTAL" in desc.upper(): break
                
                # Calcluate average monthly (cols 2 to 13 -> C to N)
                # Filter None
                vals = [v for v in row[2:14] if isinstance(v, (int, float))]
                if vals:
                    avg = sum(vals) / len(vals)
                    if avg > 0:
                        data['income'].append({
                            "id": abs(hash(desc)) % 100000,
                            "source": desc,
                            "category": get_category(desc),
                            "amount": round(avg, 2),
                            "frequency": "monthly",
                            "startDate": "2024-01-01",
                            "endDate": ""
                        })

        # Extract Expenses (if found) - usually below income
        # If exp_start not found, search below inc_end
        # The structure showed "Expenses" at some point.
        # I'll rely on text search.
        if exp_start != -1:
             for i in range(exp_start + 1, len(rows)):
                row = rows[i]
                if not row[1]: continue
                desc = str(row[1])
                if "TOTAL" in desc.upper() or "BALANCE" in desc.upper(): break
                
                vals = [v for v in row[2:14] if isinstance(v, (int, float))]
                if vals:
                    avg = sum(vals) / len(vals)
                    if avg > 0:
                         data['expenses'].append({
                            "id": abs(hash(desc)) % 100000 + 500,
                            "description": desc,
                            "category": "other", # Default, hard to guess map
                            "amount": round(avg, 2),
                            "date": "2024-02-01",
                            "recurring": True,
                            "paymentMethod": "credit_card"
                        })

    with open('final_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    print("Exported to final_data.json")

if __name__ == "__main__":
    extract()
